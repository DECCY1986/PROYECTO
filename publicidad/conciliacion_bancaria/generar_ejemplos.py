"""
Genera archivos Excel de ejemplo para probar la herramienta de conciliación bancaria.
Ejecutar: python generar_ejemplos.py
"""
import os
try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import random

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EJEMPLOS = os.path.join(BASE_DIR, 'ejemplos')
os.makedirs(EJEMPLOS, exist_ok=True)

HEADER_FILL   = PatternFill("solid", fgColor="1f6feb")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL      = PatternFill("solid", fgColor="f0f4f8")
BORDER_THIN   = Border(bottom=Side(style='thin', color='d0d7de'))

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

def fmt_cop(n):
    return f"${n:,.0f}".replace(',', '.')

# ── Movimientos de muestra ────────────────────────────────────
random.seed(42)
start_date = date(2026, 2, 1)

conceptos_banco = [
    "TRANSFERENCIA ELECTRONICA PROVEEDOR",
    "PAGO NOMINA EMPLEADOS",
    "ABONO CLIENTE FACTURA 1023",
    "RETIRO CAJERO AUTOMATICO",
    "PAGO SERVICIOS PUBLICOS EPM",
    "TRANSFERENCIA RECIBIDA",
    "PAGO ARRIENDO OFICINA",
    "COMISION BANCARIA",
    "ABONO VENTA CONTADO",
    "PAGO IMPUESTO ICA",
    "TRANSFERENCIA A CUENTA PROPIA",
    "RECAUDO CARTERA CLIENTE",
    "PAGO PROVEEDOR SUMINISTROS",
    "DEPOSITO EN EFECTIVO",
    "NOTA DEBITO GASTOS FINANCIEROS"
]

montos_base = [1500000, 3200000, 800000, 250000, 450000,
               5600000, 1200000, 35000,  2800000, 670000,
               4500000, 1800000, 920000, 650000,  85000]

movimientos = []
saldo = 15000000
for i, (concepto, monto) in enumerate(zip(conceptos_banco, montos_base)):
    dia = start_date + timedelta(days=i * 2)
    es_credito = i % 3 != 0   # la mayoría son créditos
    debito  = 0 if es_credito else monto
    credito = monto if es_credito else 0
    saldo  += credito - debito
    movimientos.append({
        'fecha': dia,
        'descripcion': concepto,
        'referencia': f'REF{100+i:04d}',
        'debito': debito,
        'credito': credito,
        'saldo': saldo
    })

# ─────────────────────────────────────────────────────────────
#  1. BANCOLOMBIA
# ─────────────────────────────────────────────────────────────
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Movimientos"
ws1.row_dimensions[1].height = 20

# Encabezados (como los exporta Bancolombia)
headers_bc = ['Fecha', 'Descripción', 'Oficina', 'Referencia', 'Débito', 'Crédito', 'Saldo']
ws1.append(headers_bc)
style_header(ws1, 1)

for j, m in enumerate(movimientos, 2):
    ws1.append([
        m['fecha'].strftime('%d/%m/%Y'),
        m['descripcion'],
        'OFICINA CENTRO',
        m['referencia'],
        m['debito'] or '',
        m['credito'] or '',
        m['saldo']
    ])
    if j % 2 == 0:
        for cell in ws1[j]:
            cell.fill = PatternFill("solid", fgColor="f6f8fa")
    for cell in ws1[j]:
        cell.border = Border(bottom=Side(style='thin', color='e8ecf0'))

auto_width(ws1)
wb1.save(os.path.join(EJEMPLOS, 'ejemplo_bancolombia.xlsx'))
print("✅ ejemplo_bancolombia.xlsx creado")

# ─────────────────────────────────────────────────────────────
#  2. BBVA
# ─────────────────────────────────────────────────────────────
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Movimientos"

movs_bbva = movimientos[:10]  # BBVA tiene menos movimientos en el período
headers_bbva = ['Fecha Operación', 'Fecha Valor', 'Concepto', 'Importe', 'Saldo']
ws2.append(headers_bbva)
style_header(ws2, 1)

for j, m in enumerate(movs_bbva, 2):
    importe = m['credito'] if m['credito'] > 0 else -m['debito']
    ws2.append([
        m['fecha'].strftime('%d/%m/%Y'),
        m['fecha'].strftime('%d/%m/%Y'),
        m['descripcion'],
        importe,
        m['saldo']
    ])

auto_width(ws2)
wb2.save(os.path.join(EJEMPLOS, 'ejemplo_bbva.xlsx'))
print("✅ ejemplo_bbva.xlsx creado")

# ─────────────────────────────────────────────────────────────
#  3. HELISA — contiene los mismos movimientos del banco
#              MÁS 3 extras que solo están en Helisa
# ─────────────────────────────────────────────────────────────
wb3 = Workbook()
ws3 = wb3.active
ws3.title = "Movimiento Bancos"

headers_helisa = ['Fecha', 'Comprobante', 'Descripción', 'Débito', 'Crédito']
ws3.append(headers_helisa)
style_header(ws3, 1)

comprobantes = [f'CB{2600+i:04d}' for i in range(len(movimientos))]

# Mismos movimientos que el banco (conciliables)
for j, (m, comp) in enumerate(zip(movimientos, comprobantes), 2):
    ws3.append([
        m['fecha'].strftime('%d/%m/%Y'),
        comp,
        m['descripcion'],
        m['debito'] or '',
        m['credito'] or ''
    ])

# 3 registros SOLO en Helisa (cheques emitidos pendientes de pago)
extras_helisa = [
    (start_date + timedelta(days=29), 'CB9901', 'CHEQUE EMITIDO PROVEEDOR XYZ',      850000, 0),
    (start_date + timedelta(days=27), 'CB9902', 'CHEQUE EMITIDO CONTRATISTA ABC',    1200000, 0),
    (start_date + timedelta(days=25), 'CB9903', 'PAGO REGISTRADO PDTE COMPENSACION',  375000, 0),
]
for fecha, comp, desc, deb, cred in extras_helisa:
    ws3.append([fecha.strftime('%d/%m/%Y'), comp, desc, deb or '', cred or ''])

auto_width(ws3)
wb3.save(os.path.join(EJEMPLOS, 'ejemplo_helisa.xlsx'))
print("✅ ejemplo_helisa.xlsx creado")

print("\n🎉 Archivos de ejemplo generados en:", EJEMPLOS)
print("   Cargue 'ejemplo_bancolombia.xlsx' como Extracto Banco (cuenta Bancolombia #1)")
print("   Cargue 'ejemplo_helisa.xlsx' como Exportación Helisa")
print("   Debería ver 15 conciliados y 3 diferencias en Helisa")
